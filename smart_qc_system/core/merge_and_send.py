import os
import sys
import glob
import re
import shutil
from datetime import datetime
import pandas as pd
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

USER_LIST_FORMATTED = [
    "이완희 (ohlee@absfil.com)",
    "조경석 (kscho@absfil.com)",
    "김민우 (mwkim@absfil.com)",
    "이홍규 (hklee@absfil.com)",
    "생산팀 김소현 (grace@absfil.com)",
    "생산팀 오예진 (yjoh@absfil.com)",
    "생산팀 내부포장 (production4@absfil.com)"
]

def extract_email(text):
    match = re.search(r'\((.*?)\)', text)
    return match.group(1).strip() if match else text

TARGET_DIR = r"C:\smart_qc_system\data\reports\input"        
RESULT_FILE = r"C:\smart_qc_system\data\reports\output\출하검사 기록관리 대장.xlsx" 
SMTP_SERVER = "wsmtp.ecount.com"  
SMTP_PORT = 587                    
DRIVE_SHARE_LINK = "https://drive.google.com/drive/folders/10yI79ChsJ3R3GLpeGjLxBYlev6-nk7BG?usp=sharing"

DESKTOP_PATH = os.path.join(os.path.expanduser("~"), "Desktop")
DESKTOP_RESULT_FILE = os.path.join(DESKTOP_PATH, "출하검사 기록관리 대장.xlsx")

def save_and_format_excel(df, file_path):
    try:
        df.to_excel(file_path, index=False)
        wb = load_workbook(file_path)
        ws = wb.active
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.value is not None:
                    val_str = str(cell.value)
                    byte_len = len(val_str.encode('utf-8'))
                    visual_len = len(val_str) + (byte_len - len(val_str)) // 2
                    if visual_len > max_len:
                        max_len = visual_len
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        wb.save(file_path)
        wb.close()
        return True
    except Exception:
        return False

def merge_excel_files(input_dir, output_file):
    all_files = glob.glob(os.path.join(input_dir, "*.xlsx"))
    if not all_files:
        return None
    
    existing_df = pd.DataFrame()
    if os.path.exists(output_file):
        try:
            existing_df = pd.read_excel(output_file)
        except Exception:
            pass
    
    new_combined_df = pd.DataFrame()
    for file in all_files:
        try:
            df = pd.read_excel(file)
            new_combined_df = pd.concat([new_combined_df, df], ignore_index=True)
        except Exception:
            pass
            
    if new_combined_df.empty and existing_df.empty:
        return None
        
    if not new_combined_df.empty:
        current_time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        if 'No.' in new_combined_df.columns:
            new_combined_df = new_combined_df.drop(columns=['No.'])
        if '통합작성시간' in new_combined_df.columns:
            new_combined_df = new_combined_df.drop(columns=['통합작성시간'])
            
        new_combined_df.insert(0, '통합작성시간', current_time_str)
        
        if not existing_df.empty:
            if 'No.' in existing_df.columns:
                existing_df = existing_df.drop(columns=['No.'])
            final_df = pd.concat([new_combined_df, existing_df], ignore_index=True)
        else:
            final_df = new_combined_df
            
        core_cols = [col for col in ['문서번호', 'P/N (품명)', 'LOT No.'] if col in final_df.columns]
        if core_cols:
            final_df = final_df.drop_duplicates(subset=core_cols, keep='first')
    else:
        final_df = existing_df
        if 'No.' in final_df.columns:
            final_df = final_df.drop(columns=['No.'])

    final_df = final_df.reset_index(drop=True)
    final_df.insert(0, 'No.', range(1, len(final_df) + 1))
    
    save_and_format_excel(final_df, output_file)
    save_and_format_excel(final_df, DESKTOP_RESULT_FILE)
    
    return final_df

def send_ecount_html_email(smtp_server, smtp_port, sender_email, sender_password, receiver_emails, today_df, available_cols):
    now = datetime.now()
    today_str = now.strftime('%Y-%m-%d')
    current_full_time = now.strftime('%Y-%m-%d %H:%M:%S')
    
    msg = MIMEMultipart('alternative')
    msg['From'] = sender_email
    msg['To'] = ", ".join(receiver_emails)
    msg['Subject'] = f"[자동발송] {current_full_time} 출하검사 기록관리 보고"

    table_html = '<table style="border-collapse: collapse; width: 100%; max-width: 1000px; font-family: \'Malgun Gothic\', sans-serif; font-size: 13px;"><thead><tr style="background-color: #e6f0fa;">'
    for col in available_cols:
        table_html += f'<th style="border: 2px solid #0056b3; padding: 12px 16px; font-weight: bold; text-align: center; color: #003366; white-space: nowrap;">{col}</th>'
    msg_table_end = '</tr></thead><tbody>'
    table_html += msg_table_end
    for i, row in today_df[available_cols].reset_index(drop=True).iterrows():
        bg_style = 'background-color: #f7faff;' if i % 2 == 0 else ''
        table_html += f'<tr style="{bg_style}">'
        for val in row:
            val_str = str(val) if pd.notnull(val) else ''
            table_html += f'<td style="border: 1px solid #0056b3; padding: 10px 14px; text-align: center; white-space: nowrap;">{val_str}</td>'
        table_html += '</tr>'
    table_html += '</tbody></table>'

    html_body = f"""<html><body><p style="font-family: 'Malgun Gothic'; font-size: 14px; font-weight: bold;">안녕하세요,</p>
    <p style="font-family: 'Malgun Gothic'; font-size: 14px;">금일 {today_str} 자로 자동 취합된 출하검사 기록관리 보고입니다.</p>
    <p style="font-family: 'Malgun Gothic'; font-size: 14px; color: #0056b3; font-weight: bold;">금일 신규 추가 {len(today_df)}건 입니다.</p><br>{table_html}
    <p style="font-family: 'Malgun Gothic'; font-size: 14px; margin-top: 15px;"><b>🔗 공유 구글 드라이브 바로가기:</b> <a href="{DRIVE_SHARE_LINK}">[absfil_uploads 폴더 열기]</a></p><br>
    <p style="font-family: 'Malgun Gothic'; font-size: 12px; color: #666666;">
        ※ 상세 원본 파일은 
        <a href="file:///C:/smart_qc_system/data/reports/output" style="color: #666666; text-decoration: underline;">
            <b>C:\smart_qc_system\data\reports\output</b>
        </a> 
        폴더에 저장되었습니다.
    </p></body></html>"""
    msg.attach(MIMEText(html_body, 'html', 'utf-8'))
    
    try:
        server = smtplib.SMTP(smtp_server, int(smtp_port), timeout=15)
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, receiver_emails, msg.as_string())
        server.quit()
        return True
    except Exception:
        return False

try:
    import streamlit as strl
    is_streamlit_mode = strl.runtime.exists() if hasattr(strl, 'runtime') else False
except ImportError:
    is_streamlit_mode = False

if is_streamlit_mode:
    strl.title("📋 출하검사 기록관리 시스템")
    strl.markdown("데이터를 검증 및 업로드하고 메일을 발송하는 통합 대시보드입니다.")

    if 'today_df_cached' not in strl.session_state:
        strl.session_state.today_df_cached = None
    if 'available_cols_cached' not in strl.session_state:
        strl.session_state.available_cols_cached = None

    strl.markdown("---")
    strl.markdown("### 🚀 1단계: 엑셀 파일 취합")
    if strl.button("엑셀 취합 및 정렬 시작", type="primary"):
        merged_data = merge_excel_files(TARGET_DIR, RESULT_FILE)
        if merged_data is not None:
            today_str = datetime.now().strftime('%Y-%m-%d')
            merged_data['통합작성시간'] = merged_data['통합작성시간'].astype(str)
            today_df = merged_data[merged_data['통합작성시간'].str.contains(today_str)].copy()
            if today_df.empty:
                strl.warning("⚠️ 파일 취합은 성공했으나, 오늘 날짜로 새롭게 추가된 신규 출하검사 데이터가 없습니다.")
            else:
                today_df = today_df.reset_index(drop=True)
                today_df['No.'] = range(1, len(today_df) + 1)
                target_columns = ['No.', '통합작성시간', '문서번호', '검사일자', '검사자', 'P/N (품명)', 'LOT No.']
                available_cols = [col for col in target_columns if col in today_df.columns]
                strl.session_state.today_df_cached = today_df
                strl.session_state.available_cols_cached = available_cols
                strl.success(f"📊 오늘자 보고 데이터 필터링 완료 및 바탕화면 저장 성공! (총 {len(today_df)}건 추가됨)")

    if strl.session_state.today_df_cached is not None:
        strl.markdown("#### [취합된 오늘 자 신규 데이터 내역]")
        strl.dataframe(strl.session_state.today_df_cached[strl.session_state.available_cols_cached], width='stretch', hide_index=True)

    strl.write("---")
    strl.markdown("### 📥 2단계: 수동 파일 업로드")
    strl.info(f"💡 취합을 진행하면 바탕화면에 **[출하검사 기록관리 대장.xlsx]** 파일이 실시간 동기화 복사 저장됩니다. 아래 버튼을 눌러 공유 폴더에 수동으로 업로드(드래그)해 주세요.")
    strl.link_button("📂 구글 드라이브 공유 폴더 열기 (absfil_uploads)", DRIVE_SHARE_LINK)
    
    strl.write("---")
    strl.markdown("### ✉️ 3단계: 최종 메일 발송 설정")
    
    # 1. 발신인 전용 리스트 정의 (USER_LIST_FORMATTED의 앞 4명만 추출)
    SENDER_LIST = USER_LIST_FORMATTED[:4]
    
    # 2. 발신인 선택 박스 (기본값 조경석은 index 1번)
    selected_sender_raw = strl.selectbox("보내는 사람 (발신인 단일 선택)", options=SENDER_LIST, index=1)
    SENDER_ID = extract_email(selected_sender_raw)
    
    SENDER_PW = strl.text_input("🔑 본인 메일 비밀번호 입력", type="password", placeholder="비밀번호를 입력하세요.")
    
    # 3. 수신인 선택 박스 (기존 USER_LIST_FORMATTED 전체 사용)
    selected_receivers_raw = strl.multiselect("받는 사람 (수신인 복수 선택 가능)", options=USER_LIST_FORMATTED)
    RECEIVER_IDS = [extract_email(item) for item in selected_receivers_raw]
    
    strl.write("") 
    
    # 🛠️ [들여쓰기 영구 박멸 마감] 중첩 구조를 100% 철거하고 단일 깊이 독립 제어식(Flat Guard Sequence) 완벽 적용
    if strl.button("✉️ 최종 보고 메일 발송하기", type="secondary"):
        if strl.session_state.today_df_cached is None:
            strl.error("❌ 먼저 상단의 '1단계: 엑셀 취합 및 정렬 시작' 버튼을 눌러 데이터를 취합하셔야 합니다.")
        elif not SENDER_PW:
            strl.error("❌ 메일을 발송하려면 비밀번호를 반드시 입력해야 합니다.")
        elif not RECEIVER_IDS:
            strl.error("❌ 수신인을 최소 1명 이상 선택해 주세요.")
        else:
            success_status = send_ecount_html_email(SMTP_SERVER, SMTP_PORT, SENDER_ID, SENDER_PW, RECEIVER_IDS, strl.session_state.today_df_cached, strl.session_state.available_cols_cached)
            if success_status:
                strl.success("📧 이카운트 메일 서버로 보고 대장을 성공적으로 전송 완료했습니다!")
